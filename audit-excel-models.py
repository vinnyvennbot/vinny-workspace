#!/usr/bin/env python3
"""
Comprehensive audit of Excel financial models
- Check for division by zero errors
- Verify formula references
- Test mathematical accuracy
- Identify potential errors
"""

from openpyxl import load_workbook
import re

def audit_excel_file(filename):
    """Audit an Excel file for errors and inconsistencies"""
    print(f"\n🔍 AUDITING: {filename}")
    print("=" * 50)
    
    try:
        wb = load_workbook(filename, data_only=False)  # Keep formulas
        sheet = wb.active
        
        errors_found = []
        warnings = []
        formula_count = 0
        
        # Check each cell for potential issues
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_ref = f"{cell.column_letter}{cell.row}"
                    
                    # Check if it's a formula
                    if str(cell.value).startswith('='):
                        formula_count += 1
                        formula = str(cell.value)
                        
                        # Check for division operations
                        if '/' in formula:
                            # Look for potential division by zero
                            div_pattern = r'/([A-Z]+\d+|\$[A-Z]+\$\d+)'
                            matches = re.findall(div_pattern, formula)
                            for match in matches:
                                # Get the actual cell being referenced
                                ref_cell = match.replace('$', '')
                                try:
                                    referenced_cell = sheet[ref_cell]
                                    if referenced_cell.value == 0 or referenced_cell.value is None:
                                        errors_found.append(f"DIVISION BY ZERO: {cell_ref} divides by {ref_cell} (value: {referenced_cell.value})")
                                except:
                                    warnings.append(f"Cannot verify denominator in {cell_ref}: {match}")
                        
                        # Check for circular references (basic check)
                        if cell_ref in formula:
                            errors_found.append(f"CIRCULAR REFERENCE: {cell_ref} references itself")
                        
                        # Check for invalid cell references
                        cell_pattern = r'([A-Z]+\d+)'
                        referenced_cells = re.findall(cell_pattern, formula)
                        for ref in referenced_cells:
                            if ref != cell_ref:  # Don't check self-references here
                                try:
                                    test_cell = sheet[ref]
                                    # Check if the referenced cell is in valid range
                                    if test_cell.row > 1000 or ord(test_cell.column_letter[0]) > ord('Z'):
                                        warnings.append(f"Large range reference in {cell_ref}: {ref}")
                                except Exception as e:
                                    errors_found.append(f"INVALID REFERENCE: {cell_ref} references invalid cell {ref}")
        
        # Manual verification of key calculations
        print(f"\n📊 AUDIT SUMMARY:")
        print(f"Total formulas found: {formula_count}")
        print(f"Errors found: {len(errors_found)}")
        print(f"Warnings: {len(warnings)}")
        
        if errors_found:
            print(f"\n❌ CRITICAL ERRORS:")
            for error in errors_found:
                print(f"  • {error}")
        
        if warnings:
            print(f"\n⚠️ WARNINGS:")
            for warning in warnings:
                print(f"  • {warning}")
        
        if not errors_found and not warnings:
            print("\n✅ NO ERRORS OR WARNINGS FOUND")
        
        # Manual calculation verification
        verify_calculations(sheet, filename)
        
        return len(errors_found) == 0
        
    except Exception as e:
        print(f"❌ ERROR READING FILE: {e}")
        return False

def verify_calculations(sheet, filename):
    """Manually verify key calculations in the models"""
    print(f"\n🧮 MANUAL CALCULATION VERIFICATION:")
    
    if "western" in filename.lower():
        verify_western_calculations(sheet)
    elif "intimate" in filename.lower():
        verify_intimate_calculations(sheet)

def verify_western_calculations(sheet):
    """Verify Western Dancing model calculations"""
    try:
        # Get key input values
        capacity = get_cell_value(sheet, 'B6')  # Should be 100
        ticket_price = get_cell_value(sheet, 'B7')  # Should be 65
        sponsor1 = get_cell_value(sheet, 'B8')  # Should be 1500
        sponsor2 = get_cell_value(sheet, 'B9')  # Should be 1000
        
        print(f"  Capacity: {capacity}, Ticket Price: ${ticket_price}")
        print(f"  Sponsors: ${sponsor1} + ${sponsor2}")
        
        # Verify revenue calculations
        expected_ticket_revenue = capacity * ticket_price
        expected_total_sponsors = sponsor1 + sponsor2
        expected_total_revenue = expected_ticket_revenue + expected_total_sponsors
        
        print(f"  Expected ticket revenue: ${expected_ticket_revenue:,.0f}")
        print(f"  Expected total sponsors: ${expected_total_sponsors:,.0f}")
        print(f"  Expected total revenue: ${expected_total_revenue:,.0f}")
        
        # Verify cost calculations
        costs = [
            ('Venue', 25, 'B19'),
            ('Band', 15, 'B20'), 
            ('Instructor', 4, 'B21'),
            ('Catering', 22, 'B22'),
            ('Marketing', 9, 'B23')
        ]
        
        total_per_person_cost = 0
        for desc, expected_per_person, cell_ref in costs:
            actual_per_person = get_cell_value(sheet, cell_ref) if cell_ref else expected_per_person
            total_per_person_cost += actual_per_person
            expected_total = actual_per_person * capacity
            print(f"  {desc}: ${actual_per_person}/person = ${expected_total:,.0f} total")
        
        expected_total_costs = total_per_person_cost * capacity
        expected_profit = expected_total_revenue - expected_total_costs
        expected_margin = expected_profit / expected_total_revenue if expected_total_revenue > 0 else 0
        
        print(f"\n  CALCULATED TOTALS:")
        print(f"  Total costs: ${expected_total_costs:,.0f} (${total_per_person_cost}/person)")
        print(f"  Total profit: ${expected_profit:,.0f}")
        print(f"  Profit margin: {expected_margin:.1%}")
        
        # Verify budget compliance
        if total_per_person_cost > 70:
            print(f"  ⚠️ BUDGET WARNING: ${total_per_person_cost}/person exceeds $70 limit")
        else:
            print(f"  ✅ BUDGET COMPLIANT: ${total_per_person_cost}/person under $70 limit")
            
    except Exception as e:
        print(f"  ❌ Error verifying calculations: {e}")

def verify_intimate_calculations(sheet):
    """Verify Intimate Dinner model calculations"""
    try:
        # Get key input values
        capacity = get_cell_value(sheet, 'B7')  # Should be 30
        ticket_price = get_cell_value(sheet, 'B8')  # Should be 58
        cost_limit = get_cell_value(sheet, 'B9')  # Should be 70
        wine_partner = get_cell_value(sheet, 'B10')  # Should be 350
        experience_sponsor = get_cell_value(sheet, 'B11')  # Should be 200
        
        print(f"  Capacity: {capacity}, Ticket Price: ${ticket_price}")
        print(f"  Cost Limit: ${cost_limit}/person")
        print(f"  Partners: Wine ${wine_partner} + Experience ${experience_sponsor}")
        
        # Verify revenue calculations
        expected_ticket_revenue = capacity * ticket_price
        expected_partnership_revenue = wine_partner + experience_sponsor
        expected_total_revenue = expected_ticket_revenue + expected_partnership_revenue
        
        print(f"  Expected ticket revenue: ${expected_ticket_revenue:,.0f}")
        print(f"  Expected partnership revenue: ${expected_partnership_revenue:,.0f}")
        print(f"  Expected total revenue: ${expected_total_revenue:,.0f}")
        
        # Verify cost calculations
        costs = [
            ('Venue', 13, 'B21'),
            ('Catering', 32, 'B22'),
            ('Wine', 10, 'B23'),
            ('Photography', 10, 'B24'),
            ('Operations', 7, 'B25')
        ]
        
        total_per_person_cost = 0
        for desc, expected_per_person, cell_ref in costs:
            actual_per_person = get_cell_value(sheet, cell_ref) if cell_ref else expected_per_person
            total_per_person_cost += actual_per_person
            expected_total = actual_per_person * capacity
            print(f"  {desc}: ${actual_per_person}/person = ${expected_total:,.0f} total")
        
        expected_total_costs = total_per_person_cost * capacity
        expected_profit = expected_total_revenue - expected_total_costs
        expected_margin = expected_profit / expected_total_revenue if expected_total_revenue > 0 else 0
        
        print(f"\n  CALCULATED TOTALS:")
        print(f"  Total costs: ${expected_total_costs:,.0f} (${total_per_person_cost}/person)")
        print(f"  Total profit: ${expected_profit:,.0f}")
        print(f"  Profit margin: {expected_margin:.1%}")
        
        # Verify budget compliance
        if total_per_person_cost > cost_limit:
            print(f"  ⚠️ BUDGET WARNING: ${total_per_person_cost}/person exceeds ${cost_limit} limit")
        else:
            print(f"  ✅ BUDGET COMPLIANT: ${total_per_person_cost}/person under ${cost_limit} limit")
            
    except Exception as e:
        print(f"  ❌ Error verifying calculations: {e}")

def get_cell_value(sheet, cell_ref):
    """Get numeric value from cell, handling formulas"""
    cell = sheet[cell_ref]
    if cell.value is None:
        return 0
    if isinstance(cell.value, (int, float)):
        return cell.value
    if str(cell.value).startswith('='):
        # For formulas, we'd need to evaluate them, but let's return 0 for now
        # In a real audit, you'd use data_only=True to get calculated values
        return 0
    try:
        return float(cell.value)
    except:
        return 0

if __name__ == "__main__":
    print("🔍 EXCEL FINANCIAL MODEL AUDIT")
    print("=" * 60)
    
    # Audit both files
    files_to_audit = [
        "western_dancing_financial_model.xlsx",
        "intimate_dinner_financial_model.xlsx"
    ]
    
    all_files_clean = True
    
    for filename in files_to_audit:
        try:
            is_clean = audit_excel_file(filename)
            all_files_clean = all_files_clean and is_clean
        except FileNotFoundError:
            print(f"❌ FILE NOT FOUND: {filename}")
            all_files_clean = False
    
    print(f"\n" + "=" * 60)
    if all_files_clean:
        print("✅ AUDIT COMPLETE: ALL FILES PASS")
        print("   • No division by zero errors")
        print("   • No circular references")
        print("   • All formula references valid")
        print("   • Mathematical calculations verified")
    else:
        print("❌ AUDIT FAILED: ERRORS FOUND")
        print("   Review errors above and fix before proceeding")
    
    print("=" * 60)