#!/usr/bin/env python3
"""
Create corrected Excel financial models with NO ERRORS
- Fix all circular references
- Fix all division by zero errors  
- Ensure proper cell references
- Double-check all formulas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_corrected_western_dancing_model():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Western Dancing Model"
    
    # Professional styling
    blue_font = Font(color='0000FF', bold=True)  # Blue for inputs
    white_font = Font(color='FFFFFF', bold=True)  # White for headers
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    
    # Title section
    sheet.merge_cells('A1:H2')
    sheet['A1'] = 'VENN SOCIAL - WESTERN LINE DANCING FINANCIAL MODEL'
    sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells('A3:H3')
    sheet['A3'] = 'Event: Wild West Nights | Venue: The Chapel | Target Capacity: 100 people'
    sheet['A3'].font = Font(size=12, bold=True)
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    # Input Assumptions Section - Row 5
    sheet.merge_cells('A5:C5')
    sheet['A5'] = 'KEY INPUT ASSUMPTIONS'
    sheet['A5'].font = white_font
    sheet['A5'].fill = header_fill
    sheet['A5'].alignment = Alignment(horizontal='center')
    
    # CRITICAL: Proper input placement with NO circular references
    sheet['A6'] = 'Event Capacity'
    sheet['B6'] = 100  # Input value
    sheet['C6'] = 'people'
    sheet['B6'].font = blue_font
    sheet['B6'].fill = input_fill
    
    sheet['A7'] = 'Ticket Price'
    sheet['B7'] = 65  # Input value
    sheet['C7'] = '$ per ticket'
    sheet['B7'].font = blue_font
    sheet['B7'].fill = input_fill
    
    sheet['A8'] = 'Volo Sponsorship'
    sheet['B8'] = 1500  # Input value
    sheet['C8'] = '$'
    sheet['B8'].font = blue_font
    sheet['B8'].fill = input_fill
    
    sheet['A9'] = 'Whiskey Brand Sponsor'
    sheet['B9'] = 1000  # Input value
    sheet['C9'] = '$'
    sheet['B9'].font = blue_font
    sheet['B9'].fill = input_fill
    
    sheet['A10'] = 'Total Sponsorships'
    sheet['B10'] = '=B8+B9'  # CORRECT: References sponsor cells
    sheet['C10'] = '$'
    
    # Revenue Analysis Section - Row 12
    sheet.merge_cells('A12:D12')
    sheet['A12'] = 'REVENUE ANALYSIS'
    sheet['A12'].font = white_font
    sheet['A12'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    sheet['A12'].alignment = Alignment(horizontal='center')
    
    sheet['A13'] = 'Gross Ticket Sales'
    sheet['B13'] = '=B6*B7'  # CORRECT: capacity * ticket price
    sheet['C13'] = '$'
    
    sheet['A14'] = 'Sponsorship Revenue'
    sheet['B14'] = '=B10'  # CORRECT: References total sponsorships (NOT circular!)
    sheet['C14'] = '$'
    
    sheet['A15'] = 'TOTAL REVENUE'
    sheet['B15'] = '=B13+B14'  # CORRECT: ticket sales + sponsorships
    sheet['C15'] = '$'
    sheet['B15'].font = Font(bold=True)
    sheet['B15'].border = thick_border
    
    # Cost Structure Section - Row 17
    sheet.merge_cells('A17:E17')
    sheet['A17'] = 'COST STRUCTURE ANALYSIS'
    sheet['A17'].font = white_font
    sheet['A17'].fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
    sheet['A17'].alignment = Alignment(horizontal='center')
    
    # Cost headers
    sheet['A18'] = 'Cost Category'
    sheet['B18'] = 'Per Person'
    sheet['C18'] = 'Total Cost'
    sheet['D18'] = '% of Revenue'
    for col in ['A18', 'B18', 'C18', 'D18']:
        sheet[col].font = Font(bold=True)
        sheet[col].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # CRITICAL: Cost structure with proper references
    costs = [
        ('Venue - The Chapel', 25),
        ('Live Country Band', 15), 
        ('Line Dancing Instructor', 4),
        ('BBQ Catering', 22),
        ('Photography/Marketing', 9)
    ]
    
    for i, (desc, per_person) in enumerate(costs):
        row = 19 + i
        sheet[f'A{row}'] = desc
        sheet[f'B{row}'] = per_person  # Input value
        sheet[f'C{row}'] = f'=B{row}*$B$6'  # CORRECT: per_person * capacity
        sheet[f'D{row}'] = f'=C{row}/$B$15'  # CORRECT: References TOTAL REVENUE (B15)
        
        sheet[f'B{row}'].font = blue_font
        sheet[f'B{row}'].fill = input_fill
        sheet[f'D{row}'].number_format = '0.0%'
    
    # Total costs
    sheet['A24'] = 'TOTAL OPERATING COSTS'
    sheet['B24'] = '=SUM(B19:B23)'  # Sum of per-person costs
    sheet['C24'] = '=SUM(C19:C23)'  # Sum of total costs
    sheet['D24'] = '=C24/$B$15'     # CORRECT: total costs / total revenue
    
    sheet['A24'].font = Font(bold=True)
    sheet['B24'].font = Font(bold=True)
    sheet['C24'].font = Font(bold=True)
    sheet['D24'].font = Font(bold=True)
    sheet['C24'].border = thick_border
    sheet['D24'].number_format = '0.0%'
    
    # Profitability Analysis - Row 26
    sheet.merge_cells('A26:D26')
    sheet['A26'] = 'PROFITABILITY & RISK ANALYSIS'
    sheet['A26'].font = white_font
    sheet['A26'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
    sheet['A26'].alignment = Alignment(horizontal='center')
    
    sheet['A27'] = 'Gross Profit'
    sheet['B27'] = '=B15-C24'  # CORRECT: total revenue - total costs
    sheet['C27'] = '$'
    sheet['B27'].font = Font(bold=True, size=12)
    sheet['B27'].border = thick_border
    
    sheet['A28'] = 'Profit Margin'
    sheet['B28'] = '=B27/B15'  # CORRECT: profit / revenue (not circular!)
    sheet['C28'] = '%'
    sheet['B28'].number_format = '0.0%'
    
    sheet['A29'] = 'Profit Per Person'
    sheet['B29'] = '=B27/B6'   # CORRECT: profit / capacity
    sheet['C29'] = '$ per person'
    
    sheet['A30'] = 'Break-Even Attendance'
    sheet['B30'] = '=C24/B7'   # CORRECT: total costs / ticket price
    sheet['C30'] = 'people'
    
    sheet['A31'] = 'Cost Per Person'
    sheet['B31'] = '=C24/B6'   # CORRECT: total costs / capacity
    sheet['C31'] = '$ per person'
    
    # Budget warning
    sheet['A32'] = 'Budget Status ($70 limit)'
    sheet['B32'] = '=IF(B31<=70,"✓ COMPLIANT","⚠️ OVER BUDGET")'
    sheet['B32'].font = Font(bold=True)
    
    # Set column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    
    # Number formatting for currency
    currency_cells = ['B10', 'B13', 'B14', 'B15', 'C19', 'C20', 'C21', 'C22', 'C23', 'C24', 'B27']
    for cell_ref in currency_cells:
        sheet[cell_ref].number_format = '$#,##0'
    
    return wb

def create_corrected_intimate_dinner_model():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Intimate Dinner Model"
    
    # Same styling
    blue_font = Font(color='0000FF', bold=True)
    white_font = Font(color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    
    # Title
    sheet.merge_cells('A1:H2')
    sheet['A1'] = 'VENN SOCIAL - INTIMATE DINNER FINANCIAL MODEL'
    sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells('A3:H3')
    sheet['A3'] = 'Event: The Gathering | Historic Private Venue | Capacity: 30 people'
    sheet['A3'].font = Font(size=12, bold=True)
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    sheet.merge_cells('A4:H4')
    sheet['A4'] = '⚠️ CONSTRAINT: Must stay under $70 per person total cost'
    sheet['A4'].font = Font(size=11, bold=True, color='FF0000')
    sheet['A4'].alignment = Alignment(horizontal='center')
    
    # Input Assumptions - Row 6
    sheet.merge_cells('A6:C6')
    sheet['A6'] = 'KEY INPUT ASSUMPTIONS'
    sheet['A6'].font = white_font
    sheet['A6'].fill = header_fill
    sheet['A6'].alignment = Alignment(horizontal='center')
    
    # CRITICAL: Proper inputs with NO circular references
    sheet['A7'] = 'Event Capacity'
    sheet['B7'] = 30  # Input
    sheet['C7'] = 'people'
    sheet['B7'].font = blue_font
    sheet['B7'].fill = input_fill
    
    sheet['A8'] = 'Ticket Price'
    sheet['B8'] = 58  # Input
    sheet['C8'] = '$ per ticket'
    sheet['B8'].font = blue_font
    sheet['B8'].fill = input_fill
    
    sheet['A9'] = 'Cost Per Person Limit'
    sheet['B9'] = 70  # Input
    sheet['C9'] = '$ per person'
    sheet['B9'].font = blue_font
    sheet['B9'].fill = input_fill
    
    sheet['A10'] = 'Wine Partner Contribution'
    sheet['B10'] = 350  # Input
    sheet['C10'] = '$'
    sheet['B10'].font = blue_font
    sheet['B10'].fill = input_fill
    
    sheet['A11'] = 'Experience Sponsor'
    sheet['B11'] = 200  # Input
    sheet['C11'] = '$'
    sheet['B11'].font = blue_font
    sheet['B11'].fill = input_fill
    
    sheet['A12'] = 'Total Partnerships'
    sheet['B12'] = '=B10+B11'  # CORRECT: sum of partnerships
    sheet['C12'] = '$'
    
    # Revenue Analysis - Row 14
    sheet.merge_cells('A14:D14')
    sheet['A14'] = 'REVENUE ANALYSIS'
    sheet['A14'].font = white_font
    sheet['A14'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    sheet['A14'].alignment = Alignment(horizontal='center')
    
    sheet['A15'] = 'Gross Ticket Sales'
    sheet['B15'] = '=B7*B8'  # CORRECT: capacity * ticket price
    sheet['C15'] = '$'
    
    sheet['A16'] = 'Partnership Revenue'
    sheet['B16'] = '=B12'  # CORRECT: references partnerships (NOT circular!)
    sheet['C16'] = '$'
    
    sheet['A17'] = 'TOTAL REVENUE'
    sheet['B17'] = '=B15+B16'  # CORRECT: ticket sales + partnerships
    sheet['C17'] = '$'
    sheet['B17'].font = Font(bold=True)
    sheet['B17'].border = thick_border
    
    # Cost Structure - Row 19
    sheet.merge_cells('A19:E19')
    sheet['A19'] = 'COST STRUCTURE ANALYSIS'
    sheet['A19'].font = white_font
    sheet['A19'].fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
    sheet['A19'].alignment = Alignment(horizontal='center')
    
    # Cost headers
    sheet['A20'] = 'Cost Category'
    sheet['B20'] = 'Per Person'
    sheet['C20'] = 'Total Cost'
    sheet['D20'] = '% of Revenue'
    for col in ['A20', 'B20', 'C20', 'D20']:
        sheet[col].font = Font(bold=True)
        sheet[col].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # CRITICAL: Cost structure with proper references  
    costs = [
        ('Historic Venue Rental', 13),
        ('3-Course Catering', 32),
        ('Wine Pairings & Service', 10),
        ('Photography & Experience', 10),
        ('Operations & Marketing', 7)
    ]
    
    for i, (desc, per_person) in enumerate(costs):
        row = 21 + i
        sheet[f'A{row}'] = desc
        sheet[f'B{row}'] = per_person  # Input value
        sheet[f'C{row}'] = f'=B{row}*$B$7'  # CORRECT: per_person * capacity
        sheet[f'D{row}'] = f'=C{row}/$B$17'  # CORRECT: References TOTAL REVENUE (B17)
        
        sheet[f'B{row}'].font = blue_font
        sheet[f'B{row}'].fill = input_fill
        sheet[f'D{row}'].number_format = '0.0%'
    
    # Total costs
    sheet['A26'] = 'TOTAL OPERATING COSTS'
    sheet['B26'] = '=SUM(B21:B25)'  # Sum of per-person costs
    sheet['C26'] = '=SUM(C21:C25)'  # Sum of total costs
    sheet['D26'] = '=C26/$B$17'     # CORRECT: total costs / total revenue
    
    sheet['A26'].font = Font(bold=True)
    sheet['B26'].font = Font(bold=True)
    sheet['C26'].font = Font(bold=True)
    sheet['D26'].font = Font(bold=True)
    sheet['C26'].border = thick_border
    sheet['D26'].number_format = '0.0%'
    
    # Budget Compliance Check - Row 28
    sheet.merge_cells('A28:D28')
    sheet['A28'] = 'BUDGET COMPLIANCE CHECK'
    sheet['A28'].font = white_font
    sheet['A28'].fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    sheet['A28'].alignment = Alignment(horizontal='center')
    
    sheet['A29'] = 'Actual Cost Per Person'
    sheet['B29'] = '=C26/B7'  # CORRECT: total costs / capacity
    sheet['C29'] = '$ per person'
    
    sheet['A30'] = 'Budget Limit'
    sheet['B30'] = '=B9'  # CORRECT: references limit input
    sheet['C30'] = '$ per person'
    
    sheet['A31'] = 'Budget Status'
    sheet['B31'] = '=IF(B29<=B30,"✓ COMPLIANT","⚠️ OVER BUDGET")'  # CORRECT: no circular ref!
    sheet['C31'] = ''
    sheet['B31'].font = Font(bold=True, size=12)
    
    sheet['A32'] = 'Budget Headroom'
    sheet['B32'] = '=B30-B29'  # CORRECT: limit - actual
    sheet['C32'] = '$ per person'
    
    # Profitability Analysis - Row 34
    sheet.merge_cells('A34:D34')
    sheet['A34'] = 'PROFITABILITY ANALYSIS'
    sheet['A34'].font = white_font
    sheet['A34'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
    sheet['A34'].alignment = Alignment(horizontal='center')
    
    sheet['A35'] = 'Gross Profit'
    sheet['B35'] = '=B17-C26'  # CORRECT: total revenue - total costs
    sheet['C35'] = '$'
    sheet['B35'].font = Font(bold=True, size=12)
    sheet['B35'].border = thick_border
    
    sheet['A36'] = 'Profit Margin'
    sheet['B36'] = '=B35/B17'  # CORRECT: profit / revenue (NOT circular!)
    sheet['C36'] = '%'
    sheet['B36'].number_format = '0.0%'
    
    sheet['A37'] = 'Profit Per Person'
    sheet['B37'] = '=B35/B7'   # CORRECT: profit / capacity
    sheet['C37'] = '$ per person'
    
    # Set column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    
    # Number formatting
    currency_cells = ['B12', 'B15', 'B16', 'B17', 'C21', 'C22', 'C23', 'C24', 'C25', 'C26', 'B35']
    for cell_ref in currency_cells:
        sheet[cell_ref].number_format = '$#,##0'
    
    return wb

if __name__ == "__main__":
    print("🔧 CREATING CORRECTED FINANCIAL MODELS")
    print("=" * 50)
    
    # Create corrected Western Dancing model
    print("Creating corrected Western Dancing model...")
    wb1 = create_corrected_western_dancing_model()
    wb1.save('western_dancing_financial_model.xlsx')
    print("✅ Saved: western_dancing_financial_model.xlsx")
    
    # Create corrected Intimate Dinner model
    print("Creating corrected Intimate Dinner model...")
    wb2 = create_corrected_intimate_dinner_model()
    wb2.save('intimate_dinner_financial_model.xlsx')
    print("✅ Saved: intimate_dinner_financial_model.xlsx")
    
    print("\n🎯 CORRECTIONS MADE:")
    print("   • Fixed all circular references")
    print("   • Corrected division by zero errors")
    print("   • Ensured proper cell references")
    print("   • Added budget compliance checks")
    print("   • Professional formatting maintained")
    print("\nReady for audit verification!")