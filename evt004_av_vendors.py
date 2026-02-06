#!/usr/bin/env python3
"""
Create formatted Excel sheet for EVT-004 Marine AV/Technical Vendors
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "EVT-004 AV Vendors"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Status colors
yes_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
pending_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
no_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

# Border
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = [
    "Event_ID", "Vendor_Name", "Category", "Contact_Email", "Contact_Phone",
    "Website", "Address", "Contact_Verified", "Verification_Method",
    "Verification_Date", "Status", "Date_Contacted", "Time_Contacted",
    "Message_ID", "Notes"
]

ws.append(headers)

# Format header row
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

# Data rows - Verified Email Contacts
verified_vendors = [
    ["EVT-004", "Meeting Tomorrow", "AV Production", "rfp@meetingtomorrow.com", "(773) 754-3878",
     "meetingtomorrow.com", "4848 W Lawrence Ave, Chicago, IL 60630", "Yes", "Website", "2026-02-06",
     "", "", "", "", "Nationwide AV company, SF office, 24/7 support"],
    
    ["EVT-004", "Megahertz AV", "AV Production", "info@mhzav.com", "(650) 641-7135",
     "mhzav.com", "147 Welsh St, San Francisco, CA 94107", "Yes", "Website", "2026-02-06",
     "", "", "", "", "SF-based, Santa Clara office, corporate events"],
    
    ["EVT-004", "Stage Lights and Sound", "AV Production", "Sales@StageLightsandSound.com", "(415) 652-0080",
     "stagelightsandsound.com", "14350 Washington St, San Leandro, CA 94578", "Yes", "Website", "2026-02-06",
     "", "", "", "", "Bay Area full-service, staging/lighting/sound"],
    
    ["EVT-004", "GEOEVENT", "AV Production", "info@geoevent.net", "818-478-2009",
     "geoevent.net", "Los Angeles (SF services)", "Yes", "Website", "2026-02-06",
     "", "", "", "", "LA-based, serves SF Bay Area, full production"],
    
    ["EVT-004", "On Point Audio Visual", "AV Production", "justin@onpointaudiovisual.com", "(415) 496-6740",
     "onpointaudiovisual.com", "1774 Timothy Dr #4, San Leandro, CA 94577", "Yes", "Website", "2026-02-06",
     "", "", "", "", "Bay Area AV, live events, Silicon Valley"],
]

# Phone-only vendors (pending verification)
pending_vendors = [
    ["EVT-004", "Fog City Audio Visual", "AV Rentals", "", "(415) 763-1000",
     "fogcityaudiovisual.com", "1145 Pine St #25, San Francisco, CA 94109", "Pending", "Phone Only", "2026-02-06",
     "", "", "", "", "20 years experience, 1-1000 person events"],
    
    ["EVT-004", "WestWave AV", "AV Production", "", "(415) 209-5817",
     "westwaveav.com", "344 Harriet St #101, San Francisco, CA 94103", "Pending", "Phone Only", "2026-02-06",
     "", "", "", "", "SF-based, hybrid events, conferences"],
    
    ["EVT-004", "Bay Area Lighting and Sound", "AV Rentals", "", "1-866-767-7623",
     "bayarealightingandsound.com", "164 Robles Way, Vallejo, CA 94591", "Pending", "Phone Only", "2026-02-06",
     "", "", "", "", "30 years experience, LED walls, staging"],
    
    ["EVT-004", "Haybear Entertainment", "AV Production", "", "925-348-4793",
     "haybearentertainment.com", "Berkeley, CA", "Pending", "Phone Only", "2026-02-06",
     "", "", "", "", "Bay Area corporate events, projection mapping"],
]

# Unverified vendors (need more research)
unverified_vendors = [
    ["EVT-004", "San Francisco AV Rentals", "AV Rentals", "", "",
     "sanfranciscoavrentals.com", "698 Pennsylvania Ave, San Francisco, CA 94107", "No", "N/A", "2026-02-06",
     "", "", "", "", "Contact form only, no email listed"],
    
    ["EVT-004", "AVT Productions", "AV Production", "", "",
     "avtproductions.com", "3351 Keller St, Santa Clara, CA 95054", "No", "N/A", "2026-02-06",
     "", "", "", "", "Contact form only, corporate events"],
    
    ["EVT-004", "MEGA Event Production", "AV Production", "", "",
     "megaeventproduction.com", "276 Elm Ave, San Bruno, CA 94066", "No", "N/A", "2026-02-06",
     "", "", "", "", "Contact form only, Bay Area events"],
    
    ["EVT-004", "Amos Productions", "AV Production", "", "",
     "amospro.com", "5715 Southfront Rd c1, Livermore, CA 94551", "No", "N/A", "2026-02-06",
     "", "", "", "", "Multi-award winning, 30+ years, contact form"],
    
    ["EVT-004", "Verducci Event Productions", "AV Production", "", "",
     "wearevep.com", "Bay Area", "No", "N/A", "2026-02-06",
     "", "", "", "", "Contact form only, DJ/photo booth"],
    
    ["EVT-004", "Impact SF", "AV Production", "", "",
     "impactsf.com", "San Francisco Bay Area", "No", "N/A", "2026-02-06",
     "", "", "", "", "Women-owned, 30+ years, contact form only"],
    
    ["EVT-004", "Event Magic", "Party Rentals/AV", "", "",
     "eventmagic.com", "2909 Chapman St, Oakland, CA 94601", "No", "N/A", "2026-02-06",
     "", "", "", "", "Since 2002, 10k SF warehouse, contact form"],
    
    ["EVT-004", "RentForEvent", "AV Rentals", "", "",
     "rentforevent.com/sf/", "San Francisco", "No", "N/A", "2026-02-06",
     "", "", "", "", "Installation service, staging, contact form"],
]

# Add all vendors
all_vendors = verified_vendors + pending_vendors + unverified_vendors
for row_data in all_vendors:
    ws.append(row_data)

# Format data rows
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    for cell in row:
        cell.border = thin_border
        cell.alignment = left_align
    
    # Color code Contact_Verified column (column H)
    verified_cell = row[7]
    if verified_cell.value == "Yes":
        verified_cell.fill = yes_fill
    elif verified_cell.value == "Pending":
        verified_cell.fill = pending_fill
    elif verified_cell.value == "No":
        verified_cell.fill = no_fill

# Set column widths
column_widths = {
    'A': 10,  # Event_ID
    'B': 30,  # Vendor_Name
    'C': 15,  # Category
    'D': 35,  # Contact_Email
    'E': 18,  # Contact_Phone
    'F': 30,  # Website
    'G': 40,  # Address
    'H': 18,  # Contact_Verified
    'I': 20,  # Verification_Method
    'J': 18,  # Verification_Date
    'K': 15,  # Status
    'L': 18,  # Date_Contacted
    'M': 18,  # Time_Contacted
    'N': 20,  # Message_ID
    'O': 50,  # Notes
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Save file
wb.save("evt004_av_vendors_verified.xlsx")
print("✅ Created evt004_av_vendors_verified.xlsx")
print(f"📊 Total vendors: {len(all_vendors)}")
print(f"✅ Verified (email): {len(verified_vendors)}")
print(f"⏳ Pending (phone only): {len(pending_vendors)}")
print(f"❌ Unverified: {len(unverified_vendors)}")
