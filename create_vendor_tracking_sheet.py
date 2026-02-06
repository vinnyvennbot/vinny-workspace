#!/usr/bin/env python3
"""Create professionally formatted vendor tracking sheet for murder mystery yacht party"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Vendor Outreach Tracker"

# Define color scheme
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
status_sent = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
status_response = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
status_quote = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

# Define border
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = [
    "Category",
    "Vendor Name", 
    "Contact Email",
    "Phone",
    "Email Sent",
    "Status",
    "Response Date",
    "Est. Cost",
    "Quote Amount",
    "Priority",
    "Notes"
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Data rows - emails sent today
sent_time = "2026-02-05 23:25 PST"
vendors_contacted = [
    ["Yacht Venue", "City Experiences", "events@cityexperiences.com", "(415) 788-8866", sent_time, "Email Sent", "", "$56K-$104K", "", "PRIMARY", "San Francisco Spirit yacht - 400 capacity. Used for Boots on Deck."],
    ["Murder Mystery", "The Murder Mystery Company", "info@murdermysteryco.com", "Via website", sent_time, "Email Sent", "", "$5K-$12K", "", "PRIMARY", "Nation's leader. 8-12 actors, scales to large venues."],
    ["AV Production", "AVT Productions", "info@avtproductions.com", "Via website", sent_time, "Email Sent", "", "$8K-$15K", "", "PRIMARY", "Bay Area based. Corporate event specialists."],
    ["Murder Mystery", "The Dinner Detective", "info@thedinnerdetective.com", "Via website", sent_time, "Email Sent", "", "$4.5K-$10K", "", "SECONDARY", "True crime approach. Located at The Marker Hotel."],
    ["AV Production", "MEGA Event Production", "info@megaeventproduction.com", "Via website", sent_time, "Email Sent", "", "$6K-$12K", "", "SECONDARY", "SF Bay Area. Experience with special events."],
    ["Murder Mystery", "Epic Immersive", "info@epicimmersive.com", "Via website", sent_time, "Email Sent", "", "$20K-$50K", "", "PREMIUM", "Unparalleled production. 100+ actors capable. 6-9 month lead."],
]

# Add vendors not yet contacted
vendors_pending = [
    ["Yacht Venue", "Empress Events", "info@empressevents.com", "Via website", "", "Research Only", "", "$150-$200/person", "", "BACKUP", "The Empress - 150 guests only. Would need 3 yachts."],
    ["Yacht Venue", "Luxe Cruises", "Via website form", "", "", "Research Only", "", "Unknown", "", "BACKUP", "Luxury brand. Need to confirm 400+ capacity."],
    ["AV Production", "Stage Lights and Sound", "info@stagelightsandsound.com", "Via website", "", "Next Wave", "", "$7.5K-$14K", "", "SECONDARY", "Good for yacht events. Marine/outdoor experience."],
    ["Event Coordinator", "TBD", "TBD", "", "", "To Be Sourced", "", "$2K-$5K", "", "REQUIRED", "Day-of coordination, vendor management."],
    ["Security", "TBD", "TBD", "", "", "To Be Sourced", "", "$1.2K-$2K", "", "REQUIRED", "4 guards for 400 guests. Crowd management."],
    ["Paramedic", "TBD", "TBD", "", "", "To Be Sourced", "", "$500-$1K", "", "RECOMMENDED", "On-site medical support."],
    ["Insurance", "TBD", "TBD", "", "", "To Be Sourced", "", "$800-$1.5K", "", "REQUIRED", "Event liability coverage."],
]

# Write data rows
row_num = 2
for vendor in vendors_contacted:
    for col_num, value in enumerate(vendor, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Color code status column
        if col_num == 6 and value == "Email Sent":
            cell.fill = status_sent
    row_num += 1

for vendor in vendors_pending:
    for col_num, value in enumerate(vendor, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row_num += 1

# Adjust column widths
column_widths = {
    'A': 18,  # Category
    'B': 25,  # Vendor Name
    'C': 28,  # Contact Email
    'D': 18,  # Phone
    'E': 22,  # Email Sent
    'F': 15,  # Status
    'G': 18,  # Response Date
    'H': 15,  # Est. Cost
    'I': 15,  # Quote Amount
    'J': 12,  # Priority
    'K': 45,  # Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Set row heights
ws.row_dimensions[1].height = 25  # Header row
for row in range(2, row_num):
    ws.row_dimensions[row].height = 35  # Data rows with wrap text

# Freeze header row
ws.freeze_panes = 'A2'

# Add summary section at top (above frozen header)
ws.insert_rows(1, 3)

# Title
title_cell = ws['A1']
title_cell.value = "Murder Mystery Yacht Party - Vendor Outreach Tracker"
title_cell.font = Font(size=16, bold=True, color="1F4E78")
ws.merge_cells('A1:K1')

# Summary stats
ws['A2'] = "Event Date:"
ws['B2'] = "Summer 2026 (TBD)"
ws['D2'] = "Capacity:"
ws['E2'] = "400 guests"
ws['G2'] = "Budget:"
ws['H2'] = "$85K-$140K total"

# Stats row
ws['A3'] = "Emails Sent:"
ws['B3'] = "6"
ws['D3'] = "Awaiting Response:"
ws['E3'] = "6"
ws['G3'] = "Last Updated:"
ws['H3'] = datetime.now().strftime("%Y-%m-%d %H:%M PST")

# Adjust freeze panes
ws.freeze_panes = 'A5'

# Save file
output_path = "/Users/vinnyvenn/.openclaw/workspace/murder-mystery-vendor-tracker.xlsx"
wb.save(output_path)

print(f"✅ Created: {output_path}")
print(f"📊 Tracking {len(vendors_contacted)} contacted vendors + {len(vendors_pending)} pending")
