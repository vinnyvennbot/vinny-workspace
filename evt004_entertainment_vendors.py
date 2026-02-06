#!/usr/bin/env python3
"""Create professionally formatted EVT-004 Entertainment Vendors spreadsheet"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "EVT-004 Entertainment"

# Define colors
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
pending_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = [
    "Event_ID", "Vendor_Name", "Category", "Contact_Email", "Contact_Phone",
    "Website", "Contact_Verified", "Verification_Method", "Verification_Date",
    "Status", "Date_Contacted", "Message_ID", "Notes"
]

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Set column widths
column_widths = {
    'A': 10,  # Event_ID
    'B': 25,  # Vendor_Name
    'C': 20,  # Category
    'D': 30,  # Contact_Email
    'E': 18,  # Contact_Phone
    'F': 25,  # Website
    'G': 18,  # Contact_Verified
    'H': 20,  # Verification_Method
    'I': 18,  # Verification_Date
    'J': 15,  # Status
    'K': 18,  # Date_Contacted
    'L': 25,  # Message_ID
    'M': 35   # Notes
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Vendor data - VERIFIED ONLY
vendors = [
    {
        "name": "Radio Gatsby",
        "category": "Jazz Band (1920s/Gatsby)",
        "email": "inquiries@radiogatsby.com",
        "phone": "415-602-3305",
        "website": "radiogatsby.com",
        "verified": "Yes",
        "method": "Website",
        "date": "2026-02-06",
        "notes": "Premier 1920s/speakeasy jazz, perfect for Gatsby theme"
    },
    {
        "name": "The Cosmo Alleycats",
        "category": "Jazz Band (Vintage/1920s)",
        "email": "inquiries@cosmoalleycats.com",
        "phone": "(415) 688-0123",
        "website": "cosmoalleycats.com",
        "verified": "Yes",
        "method": "Website",
        "date": "2026-02-06",
        "notes": "Vintage swing/jazz, speakeasy & 1920s specialists"
    },
    {
        "name": "Campbell's Jazz Soup",
        "category": "Jazz Band (1920s-50s)",
        "email": "info@campbellsjazz.com",
        "phone": "415-499-3316",
        "website": "campbellsjazz.com",
        "verified": "Yes",
        "method": "Website + Yelp",
        "date": "2026-02-06",
        "notes": "1920s-50s jazz and swing, 20+ years experience"
    },
    {
        "name": "The Harrison Jazz Ensemble",
        "category": "Jazz Band",
        "email": "",
        "phone": "(415) 573-0140",
        "website": "theharrisonjazzband.com",
        "verified": "Yes",
        "method": "Website",
        "date": "2026-02-06",
        "notes": "Professional jazz for SF/Napa events, corporate specialists"
    },
    {
        "name": "Nova Jazz Band",
        "category": "Jazz Trio/Quartet",
        "email": "",
        "phone": "(415) 519-9888",
        "website": "novajazzonline.com",
        "verified": "Yes",
        "method": "Google Places",
        "date": "2026-02-06",
        "notes": "Corporate events, elegant ambient jazz"
    },
    {
        "name": "LIV Entertainment Group",
        "category": "DJ/Live Band (Full Service)",
        "email": "",
        "phone": "(310) 699-9825",
        "website": "liventgroup.com",
        "verified": "Yes",
        "method": "Google Places",
        "date": "2026-02-06",
        "notes": "DJs, live bands, full entertainment services"
    },
    {
        "name": "Black & White Affair DJ",
        "category": "DJ (Wedding/Corporate)",
        "email": "",
        "phone": "(408) 768-4904",
        "website": "bwadj.com",
        "verified": "Yes",
        "method": "Google Places",
        "date": "2026-02-06",
        "notes": "Professional DJ services, corporate events"
    },
    {
        "name": "California Disc Jockey",
        "category": "DJ (Bay Area)",
        "email": "",
        "phone": "415-350-5474",
        "website": "californiadiscjockey.com",
        "verified": "Yes",
        "method": "Website",
        "date": "2026-02-06",
        "notes": "Chris Webb, professional Bay Area DJ since 1988"
    },
    {
        "name": "DJ Jeremy Productions",
        "category": "DJ (Wedding/Events)",
        "email": "info@djjeremyproductions.com",
        "phone": "415-964-1060",
        "website": "djjeremyproductions.com",
        "verified": "Yes",
        "method": "Website + Blog",
        "date": "2026-02-06",
        "notes": "SF Bay Area wedding and event DJs, photobooths"
    },
    {
        "name": "Subito Strings",
        "category": "String Quartet (Ambient)",
        "email": "",
        "phone": "(310) 775-7969",
        "website": "subitostrings.com",
        "verified": "Yes",
        "method": "Google Places",
        "date": "2026-02-06",
        "notes": "Conservatory-trained, pop + classical, SF/LA/SD"
    },
    {
        "name": "Bay Area Strings",
        "category": "String Quartet/Trio (Ambient)",
        "email": "info@BayAreaStrings.com",
        "phone": "415-562-8465",
        "website": "bayareastrings.com",
        "verified": "Yes",
        "method": "Website",
        "date": "2026-02-06",
        "notes": "Classical, contemporary, rock, jazz on strings"
    },
    {
        "name": "SFCM Musicians Program",
        "category": "Classical/Jazz (Full Service)",
        "email": "hiresfcmmusicians@sfcm.edu",
        "phone": "(415) 503-6297",
        "website": "sfcm.edu",
        "verified": "Yes",
        "method": "Website",
        "date": "2026-02-06",
        "notes": "SF Conservatory - strings, brass, woodwinds, jazz combos, affordable"
    },
    {
        "name": "Gatsby Event Studios",
        "category": "DJ/Entertainment (Full Service)",
        "email": "Info@gatsbyeventstudios.com",
        "phone": "(925) 200-7391",
        "website": "gatsbyeventstudios.com",
        "verified": "Yes",
        "method": "Website",
        "date": "2026-02-06",
        "notes": "A/V, DJs, entertainers - perfect name for Gatsby theme!"
    },
]

# Add vendor rows
for idx, vendor in enumerate(vendors, start=2):
    ws.cell(row=idx, column=1).value = "EVT-004"
    ws.cell(row=idx, column=2).value = vendor["name"]
    ws.cell(row=idx, column=3).value = vendor["category"]
    ws.cell(row=idx, column=4).value = vendor["email"]
    ws.cell(row=idx, column=5).value = vendor["phone"]
    ws.cell(row=idx, column=6).value = vendor["website"]
    ws.cell(row=idx, column=7).value = vendor["verified"]
    ws.cell(row=idx, column=8).value = vendor["method"]
    ws.cell(row=idx, column=9).value = vendor["date"]
    ws.cell(row=idx, column=10).value = "Researched"
    ws.cell(row=idx, column=11).value = ""
    ws.cell(row=idx, column=12).value = ""
    ws.cell(row=idx, column=13).value = vendor["notes"]
    
    # Apply borders and formatting
    for col in range(1, 14):
        cell = ws.cell(row=idx, column=col)
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Color code Contact_Verified column
    verified_cell = ws.cell(row=idx, column=7)
    if vendor["verified"] == "Yes":
        verified_cell.fill = verified_fill
    else:
        verified_cell.fill = pending_fill

# Freeze header row
ws.freeze_panes = 'A2'

# Save file
wb.save('EVT-004_Entertainment_Vendors_Verified.xlsx')
print("✅ Created: EVT-004_Entertainment_Vendors_Verified.xlsx")
print(f"📊 Total verified vendors: {len(vendors)}")
