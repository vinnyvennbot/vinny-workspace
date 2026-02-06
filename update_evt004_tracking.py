#!/usr/bin/env python3
"""Update EVT-004 vendor tracking with email send data"""

from openpyxl import load_workbook
from datetime import datetime

# Load existing workbook
wb = load_workbook("evt004_av_vendors_verified.xlsx")
ws = wb.active

# Email send data
sends = [
    {
        "row": 2,  # Meeting Tomorrow
        "status": "Email Sent",
        "date": "2026-02-06",
        "time": "00:23 PST",
        "message_id": "19c3210cd826858d"
    },
    {
        "row": 3,  # Megahertz AV
        "status": "Email Sent",
        "date": "2026-02-06",
        "time": "00:24 PST",
        "message_id": "19c3210fa442b148"
    },
    {
        "row": 4,  # Stage Lights and Sound
        "status": "Email Sent",
        "date": "2026-02-06",
        "time": "00:24 PST",
        "message_id": "19c32112fee4bcbe"
    },
    {
        "row": 5,  # GEOEVENT
        "status": "Email Sent",
        "date": "2026-02-06",
        "time": "00:24 PST",
        "message_id": "19c3211321b9bbe0"
    },
    {
        "row": 6,  # On Point Audio Visual
        "status": "Email Sent",
        "date": "2026-02-06",
        "time": "00:24 PST",
        "message_id": "19c321152ebfadf2"
    }
]

# Update cells
for send in sends:
    row = send["row"]
    ws[f"K{row}"] = send["status"]  # Status column
    ws[f"L{row}"] = send["date"]    # Date_Contacted
    ws[f"M{row}"] = send["time"]    # Time_Contacted
    ws[f"N{row}"] = send["message_id"]  # Message_ID

# Save
wb.save("evt004_av_vendors_verified.xlsx")
print("✅ Updated tracking spreadsheet with 5 email sends")
print("📧 All message IDs recorded")
