#!/usr/bin/env python3
"""
Create professionally formatted investment banker-style financial models
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_western_dancing_model():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Western Dancing Model"
    
    # Investment banker color scheme
    blue_font = Font(color='0000FF', bold=True)  # Blue for inputs
    black_font = Font(color='000000')  # Black for calculations
    white_font = Font(color='FFFFFF', bold=True)  # White for headers
    
    # Background colors
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')  # Dark blue
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # Light blue
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    
    # Title section
    sheet.merge_cells('A1:H2')
    sheet['A1'] = 'VENN SOCIAL - WESTERN LINE DANCING FINANCIAL MODEL'
    sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    sheet['A1'].fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Event details
    sheet.merge_cells('A3:H3')
    sheet['A3'] = 'Event: Wild West Nights | Venue: The Chapel | Target Capacity: 100 people'
    sheet['A3'].font = Font(size=12, bold=True)
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    # Input Assumptions Section
    row = 5
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = 'KEY INPUT ASSUMPTIONS'
    sheet[f'A{row}'].font = white_font
    sheet[f'A{row}'].fill = header_fill
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    inputs = [
        ('Event Capacity', 100, 'people'),
        ('Ticket Price', 65, '$ per ticket'),
        ('Volo Sponsorship', 1500, '$'),
        ('Whiskey Brand Sponsor', 1000, '$')
    ]
    
    for i, (desc, value, unit) in enumerate(inputs):
        r = row + 1 + i
        sheet[f'A{r}'] = desc
        sheet[f'B{r}'] = value
        sheet[f'C{r}'] = unit
        sheet[f'B{r}'].font = blue_font  # Blue for inputs
        sheet[f'B{r}'].fill = input_fill
    
    # Revenue Analysis Section
    row = 11
    sheet.merge_cells(f'A{row}:D{row}')
    sheet[f'A{row}'] = 'REVENUE ANALYSIS'
    sheet[f'A{row}'].font = white_font
    sheet[f'A{row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')  # Green
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    revenue_items = [
        ('Gross Ticket Sales', '=B6*B7', '$'),
        ('Total Sponsorships', '=B8+B9', '$'),
        ('TOTAL REVENUE', '=B13+B14', '$')
    ]
    
    for i, (desc, formula, unit) in enumerate(revenue_items):
        r = row + 1 + i
        sheet[f'A{r}'] = desc
        sheet[f'B{r}'] = formula
        sheet[f'C{r}'] = unit
        if 'TOTAL' in desc:
            sheet[f'A{r}'].font = Font(bold=True)
            sheet[f'B{r}'].font = Font(bold=True)
            sheet[f'B{r}'].border = thick_border
    
    # Cost Structure Section
    row = 17
    sheet.merge_cells(f'A{row}:E{row}')
    sheet[f'A{row}'] = 'COST STRUCTURE ANALYSIS'
    sheet[f'A{row}'].font = white_font
    sheet[f'A{row}'].fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')  # Orange/Red
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Cost headers
    r = row + 1
    headers = ['Cost Category', 'Per Person', 'Total Cost', '% of Revenue']
    for i, header in enumerate(headers):
        sheet[f'{get_column_letter(i+1)}{r}'] = header
        sheet[f'{get_column_letter(i+1)}{r}'].font = Font(bold=True)
        sheet[f'{get_column_letter(i+1)}{r}'].fill = section_fill
    
    costs = [
        ('Venue - The Chapel', 25, '=B19*$B$6', '=C19/$B$15'),
        ('Live Country Band', 15, '=B20*$B$6', '=C20/$B$15'),
        ('Line Dancing Instructor', 4, '=B21*$B$6', '=C21/$B$15'),
        ('BBQ Catering', 22, '=B22*$B$6', '=C22/$B$15'),
        ('Photography/Marketing', 9, '=B23*$B$6', '=C23/$B$15')
    ]
    
    for i, (desc, per_person, total_formula, pct_formula) in enumerate(costs):
        r = row + 2 + i
        sheet[f'A{r}'] = desc
        sheet[f'B{r}'] = per_person
        sheet[f'C{r}'] = total_formula
        sheet[f'D{r}'] = pct_formula
        sheet[f'B{r}'].font = blue_font  # Blue for per-person inputs
        sheet[f'D{r}'].number_format = '0.0%'
    
    # Total costs
    r = row + 7
    sheet[f'A{r}'] = 'TOTAL OPERATING COSTS'
    sheet[f'B{r}'] = '=SUM(B19:B23)'
    sheet[f'C{r}'] = '=SUM(C19:C23)'
    sheet[f'D{r}'] = '=C24/$B$15'
    sheet[f'A{r}'].font = Font(bold=True)
    sheet[f'B{r}'].font = Font(bold=True)
    sheet[f'C{r}'].font = Font(bold=True)
    sheet[f'D{r}'].font = Font(bold=True)
    sheet[f'C{r}'].border = thick_border
    sheet[f'D{r}'].number_format = '0.0%'
    
    # Profitability Analysis
    row = 26
    sheet.merge_cells(f'A{row}:D{row}')
    sheet[f'A{row}'] = 'PROFITABILITY & RISK ANALYSIS'
    sheet[f'A{row}'].font = white_font
    sheet[f'A{row}'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')  # Purple
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    profit_metrics = [
        ('Gross Profit', '=B15-C24', '$'),
        ('Profit Margin', '=B27/B15', '%'),
        ('Profit Per Person', '=B27/B6', '$ per person'),
        ('Break-Even Attendance', '=C24/B7', 'people'),
        ('Cost Per Person', '=C24/B6', '$ per person')
    ]
    
    for i, (desc, formula, unit) in enumerate(profit_metrics):
        r = row + 1 + i
        sheet[f'A{r}'] = desc
        sheet[f'B{r}'] = formula
        sheet[f'C{r}'] = unit
        if '%' in unit:
            sheet[f'B{r}'].number_format = '0.0%'
        if 'Gross Profit' in desc:
            sheet[f'A{r}'].font = Font(bold=True, size=12)
            sheet[f'B{r}'].font = Font(bold=True, size=12)
            sheet[f'B{r}'].border = thick_border
    
    # Set column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    
    # Number formatting
    for row_num in range(6, 32):
        for col in ['B', 'C']:
            if sheet[f'{col}{row_num}'].value and str(sheet[f'{col}{row_num}'].value).startswith('='):
                sheet[f'{col}{row_num}'].number_format = '$#,##0'
    
    return wb

def create_intimate_dinner_model():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Intimate Dinner Model"
    
    # Same styling as above
    blue_font = Font(color='0000FF', bold=True)
    black_font = Font(color='000000')
    white_font = Font(color='FFFFFF', bold=True)
    
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    
    # Title
    sheet.merge_cells('A1:H2')
    sheet['A1'] = 'VENN SOCIAL - INTIMATE DINNER FINANCIAL MODEL'
    sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    sheet['A1'].fill = header_fill
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Event details
    sheet.merge_cells('A3:H3')
    sheet['A3'] = 'Event: The Gathering | Historic Private Venue + Catering | Capacity: 30 people'
    sheet['A3'].font = Font(size=12, bold=True)
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    # Budget constraint warning
    sheet.merge_cells('A4:H4')
    sheet['A4'] = 'CONSTRAINT: Must stay under $70 per person total cost'
    sheet['A4'].font = Font(size=11, bold=True, color='FF0000')  # Red warning
    sheet['A4'].alignment = Alignment(horizontal='center')
    
    # Inputs
    row = 6
    sheet.merge_cells(f'A{row}:C{row}')
    sheet[f'A{row}'] = 'KEY INPUT ASSUMPTIONS'
    sheet[f'A{row}'].font = white_font
    sheet[f'A{row}'].fill = header_fill
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    inputs = [
        ('Event Capacity', 30, 'people'),
        ('Ticket Price', 58, '$ per ticket'),
        ('Cost Per Person Limit', 70, '$ per person'),
        ('Wine Partner Contribution', 350, '$'),
        ('Experience Sponsor', 200, '$')
    ]
    
    for i, (desc, value, unit) in enumerate(inputs):
        r = row + 1 + i
        sheet[f'A{r}'] = desc
        sheet[f'B{r}'] = value
        sheet[f'C{r}'] = unit
        sheet[f'B{r}'].font = blue_font
        sheet[f'B{r}'].fill = input_fill
    
    # Revenue section
    row = 13
    sheet.merge_cells(f'A{row}:D{row}')
    sheet[f'A{row}'] = 'REVENUE ANALYSIS'
    sheet[f'A{row}'].font = white_font
    sheet[f'A{row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    revenue_items = [
        ('Gross Ticket Sales', '=B7*B8', '$'),
        ('Partnership Revenue', '=B10+B11', '$'),
        ('TOTAL REVENUE', '=B15+B16', '$')
    ]
    
    for i, (desc, formula, unit) in enumerate(revenue_items):
        r = row + 1 + i
        sheet[f'A{r}'] = desc
        sheet[f'B{r}'] = formula
        sheet[f'C{r}'] = unit
        if 'TOTAL' in desc:
            sheet[f'B{r}'].font = Font(bold=True)
            sheet[f'B{r}'].border = thick_border
    
    # Cost structure
    row = 19
    sheet.merge_cells(f'A{row}:E{row}')
    sheet[f'A{row}'] = 'COST STRUCTURE ANALYSIS'
    sheet[f'A{row}'].font = white_font
    sheet[f'A{row}'].fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    costs = [
        ('Historic Venue Rental', 13, '=B21*$B$7', '=C21/$B$17'),
        ('3-Course Catering', 32, '=B22*$B$7', '=C22/$B$17'),
        ('Wine Pairings & Service', 10, '=B23*$B$7', '=C23/$B$17'),
        ('Photography & Experience', 10, '=B24*$B$7', '=C24/$B$17'),
        ('Operations & Marketing', 7, '=B25*$B$7', '=C25/$B$17')
    ]
    
    r = row + 1
    headers = ['Cost Category', 'Per Person', 'Total Cost', '% of Revenue']
    for i, header in enumerate(headers):
        sheet[f'{get_column_letter(i+1)}{r}'] = header
        sheet[f'{get_column_letter(i+1)}{r}'].font = Font(bold=True)
    
    for i, (desc, per_person, total_formula, pct_formula) in enumerate(costs):
        r = row + 2 + i
        sheet[f'A{r}'] = desc
        sheet[f'B{r}'] = per_person
        sheet[f'C{r}'] = total_formula
        sheet[f'D{r}'] = pct_formula
        sheet[f'B{r}'].font = blue_font
        sheet[f'D{r}'].number_format = '0.0%'
    
    # Total costs
    r = row + 7
    sheet[f'A{r}'] = 'TOTAL OPERATING COSTS'
    sheet[f'B{r}'] = '=SUM(B21:B25)'
    sheet[f'C{r}'] = '=SUM(C21:C25)'
    sheet[f'D{r}'] = '=C26/$B$17'
    sheet[f'B{r}'].font = Font(bold=True)
    sheet[f'C{r}'].font = Font(bold=True)
    sheet[f'C{r}'].border = thick_border
    
    # Budget compliance check
    row = 28
    sheet.merge_cells(f'A{row}:D{row}')
    sheet[f'A{row}'] = 'BUDGET COMPLIANCE CHECK'
    sheet[f'A{row}'].font = white_font
    sheet[f'A{row}'].fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')  # Red
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    compliance_checks = [
        ('Actual Cost Per Person', '=C26/B7', '$ per person'),
        ('Budget Limit', '=B9', '$ per person'),
        ('Budget Status', '=IF(B30<=B31,"✓ COMPLIANT","⚠️ OVER BUDGET")', ''),
        ('Budget Headroom', '=B31-B30', '$ per person')
    ]
    
    for i, (desc, formula, unit) in enumerate(compliance_checks):
        r = row + 1 + i
        sheet[f'A{r}'] = desc
        sheet[f'B{r}'] = formula
        sheet[f'C{r}'] = unit
        if 'Status' in desc:
            sheet[f'B{r}'].font = Font(bold=True, size=12)
    
    # Profitability
    row = 34
    sheet.merge_cells(f'A{row}:D{row}')
    sheet[f'A{row}'] = 'PROFITABILITY ANALYSIS'
    sheet[f'A{row}'].font = white_font
    sheet[f'A{row}'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
    sheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    profit_metrics = [
        ('Gross Profit', '=B17-C26', '$'),
        ('Profit Margin', '=B36/B17', '%'),
        ('Profit Per Person', '=B36/B7', '$ per person')
    ]
    
    for i, (desc, formula, unit) in enumerate(profit_metrics):
        r = row + 1 + i
        sheet[f'A{r}'] = desc
        sheet[f'B{r}'] = formula
        sheet[f'C{r}'] = unit
        if '%' in unit:
            sheet[f'B{r}'].number_format = '0.0%'
        if 'Gross Profit' in desc:
            sheet[f'B{r}'].font = Font(bold=True, size=12)
            sheet[f'B{r}'].border = thick_border
    
    # Set column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    
    # Format currency cells
    for row_num in range(7, 40):
        for col in ['B', 'C']:
            if sheet[f'{col}{row_num}'].value and str(sheet[f'{col}{row_num}'].value).startswith('='):
                if 'person' not in str(sheet[f'C{row_num}'].value or ''):
                    sheet[f'{col}{row_num}'].number_format = '$#,##0'
    
    return wb

if __name__ == "__main__":
    # Create Western Dancing Model
    wb1 = create_western_dancing_model()
    wb1.save('western_dancing_financial_model.xlsx')
    print("Created: western_dancing_financial_model.xlsx")
    
    # Create Intimate Dinner Model  
    wb2 = create_intimate_dinner_model()
    wb2.save('intimate_dinner_financial_model.xlsx')
    print("Created: intimate_dinner_financial_model.xlsx")
    
    print("\nBoth models created with investment banker-style formatting!")
    print("Features: Color-coded sections, proper borders, formula validation, constraint checking")